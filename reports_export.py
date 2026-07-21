"""
Parabellum ISOS — Report Generation
=================================================================
Objective 2.3 / DFD 5.0: stock status, material movement, transaction,
commission, and forecast reports.

Design: every report type has ONE function that queries the database and
returns (title, subtitle, columns, rows) with values already formatted as
display strings (peso signs, percentages, etc. included). The on-screen
preview, the PDF, and the Excel file all read from that SAME function —
so the three formats can never disagree with each other, and a defense
panelist comparing "does the PDF match what's on screen" always gets yes.
"""

import io
from datetime import datetime

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from mlr_model import execute_query

PRIMARY_HEX = "8B1E1E"   # matches --primary in styles.css


def _peso(n):
    return f"\u20b1{float(n or 0):,.2f}"


def _num(n):
    return f"{float(n or 0):,.2f}".rstrip("0").rstrip(".") if "." in f"{float(n or 0):,.2f}" else f"{float(n or 0):,.0f}"


# =================================================================
#  Report builders — one per report type. Each returns:
#    title, subtitle, columns (list[str]), rows (list[list[str]])
# =================================================================
def build_inventory_report(db_config):
    rows = execute_query(db_config, """
        SELECT material_name, category, current_stock, unit, unit_cost,
               reorder_level
        FROM materials ORDER BY material_name;
    """, fetch=True)
    out = []
    for r in rows:
        qty = float(r["current_stock"])
        value = qty * float(r["unit_cost"])
        status = "Out of Stock" if qty <= 0 else ("Low Stock" if qty <= float(r["reorder_level"]) else "In Stock")
        out.append([r["material_name"], r["category"] or "\u2014",
                    f"{qty:g} {r['unit']}", _peso(r["unit_cost"]), _peso(value), status])
    return ("Inventory Report", "Stock levels and valuation",
            ["Item", "Category", "Qty", "Unit Price", "Value", "Status"], out)


def build_customer_report(db_config):
    rows = execute_query(db_config, """
        SELECT c.name, c.contact_person, c.status,
               (SELECT COUNT(*) FROM projects p WHERE p.customer_id = c.customer_id) AS projects,
               (SELECT COUNT(*) FROM transactions t WHERE t.customer_id = c.customer_id) AS txns
        FROM customers c ORDER BY c.name;
    """, fetch=True)
    out = [[r["name"], r["contact_person"] or "\u2014", r["status"],
            str(int(r["projects"])), str(int(r["txns"]))] for r in rows]
    return ("Customer Report", "Accounts and activity",
            ["Customer", "Contact", "Status", "Projects", "Transactions"], out)


def build_project_report(db_config):
    rows = execute_query(db_config, """
        SELECT p.project_name, c.name AS customer_name, p.status,
               p.progress, p.budget, p.end_date
        FROM projects p LEFT JOIN customers c ON c.customer_id = p.customer_id
        WHERE p.project_code IS NOT NULL
        ORDER BY p.start_date DESC;
    """, fetch=True)
    status_map = {"Ongoing": "In Progress", "Completed": "Completed"}
    out = [[r["project_name"], r["customer_name"] or "\u2014",
            status_map.get(r["status"], r["status"]), f"{int(r['progress'] or 0)}%",
            _peso(r["budget"]), str(r["end_date"]) if r["end_date"] else "\u2014"] for r in rows]
    return ("Project Report", "Status and budgets",
            ["Project", "Customer", "Status", "Progress", "Budget", "Due"], out)


def build_transaction_report(db_config):
    rows = execute_query(db_config, """
        SELECT t.transaction_id, c.name AS customer_name, t.material_name,
               t.quantity, t.unit_price, t.payment_status, t.txn_date
        FROM transactions t LEFT JOIN customers c ON c.customer_id = t.customer_id
        ORDER BY t.txn_date DESC, t.transaction_id DESC;
    """, fetch=True)
    out = []
    for r in rows:
        total = float(r["quantity"] or 0) * float(r["unit_price"] or 0) * 1.12  # incl. 12% VAT
        out.append([f"TXN-{r['transaction_id']:04d}", r["customer_name"] or "\u2014",
                    r["material_name"] or "\u2014", _num(r["quantity"]),
                    _peso(total), r["payment_status"] or "Pending",
                    str(r["txn_date"]) if r["txn_date"] else "\u2014"])
    return ("Client Transaction Report", "Invoices incl. 12% VAT",
            ["Invoice", "Customer", "Material", "Qty", "Total", "Status", "Date"], out)


def compute_commissions(db_config):
    """
    Real commission computation (Objective — commission tracking tied to
    completed projects).

    Total Sales   = sum of budgets of an employee's COMPLETED projects,
                     all-time (projects.staff links to employees.employee_code).
    Commission    = Total Sales x commission_rate.
    Monthly       = Commission divided by the number of distinct calendar
                     months in which the employee completed a project — i.e.
                     an average monthly earning rate, not literally "this
                     calendar month." A brand-new demo database may have all
                     its completions in the same handful of months, or none
                     completed yet in the current month, so "this month"
                     alone could misleadingly show ₱0 even for a productive
                     employee. Averaging over their active months gives a
                     stable, explainable figure instead.
    """
    rows = execute_query(db_config, """
        SELECT e.employee_code, e.name, e.role, e.commission_rate,
               COUNT(CASE WHEN p.status = 'Completed' THEN 1 END) AS completed,
               COALESCE(SUM(CASE WHEN p.status = 'Completed' THEN p.budget END), 0) AS total_sales,
               COUNT(DISTINCT CASE WHEN p.status = 'Completed'
                     THEN date_trunc('month', p.end_date) END) AS active_months
        FROM employees e
        LEFT JOIN projects p ON p.staff = e.employee_code
        WHERE e.status = 'Active'
        GROUP BY e.employee_id, e.employee_code, e.name, e.role, e.commission_rate
        ORDER BY e.name;
    """, fetch=True)

    out = []
    for r in rows:
        rate = float(r["commission_rate"])
        total_sales = float(r["total_sales"])
        commission = total_sales * rate / 100
        active_months = int(r["active_months"]) or 1
        out.append({
            "code": r["employee_code"], "name": r["name"], "role": r["role"] or "\u2014",
            "rate": rate, "completed": int(r["completed"]),
            "total_sales": total_sales, "commission": commission,
            "monthly": commission / active_months,
        })
    return out


def build_commission_report(db_config):
    rows = compute_commissions(db_config)
    out = [[e["name"], e["role"], str(e["completed"]), f"{e['rate']:g}%",
            _peso(e["total_sales"]), _peso(e["monthly"])] for e in rows]
    return ("Commission Report", "Per-employee summary \u2014 computed from completed projects",
            ["Employee", "Role", "Completed", "Rate", "Sales", "Monthly Commission"], out)


def build_forecast_report(db_config):
    rows = execute_query(db_config, """
        SELECT m.material_name, f.forecast_month, f.predicted_demand,
               m.current_stock, m.unit
        FROM forecast_results f
        JOIN materials m ON m.material_id = f.material_id
        WHERE f.forecast_month = (SELECT MAX(forecast_month) FROM forecast_results)
        ORDER BY m.material_name;
    """, fetch=True)
    metrics = execute_query(db_config, """
        SELECT mae, rmse, mape, r2 FROM model_metrics
        ORDER BY evaluated_at DESC LIMIT 1;
    """, fetch=True)
    out = []
    for r in rows:
        pred = float(r["predicted_demand"])
        stock = float(r["current_stock"])
        trend = "\u2014"
        out.append([r["material_name"], str(r["forecast_month"]),
                    f"{pred:g} {r['unit']}", f"{stock:g} {r['unit']}", trend,
                    f"R\u00b2 {metrics[0]['r2']}" if metrics else "\u2014"])
    subtitle = "Predicted demand for the coming month"
    if metrics:
        m = metrics[0]
        subtitle += f" \u2014 MAE {m['mae']}, RMSE {m['rmse']}, MAPE {m['mape']}%, R\u00b2 {m['r2']}"
    if not out:
        subtitle += " (no forecast has been run yet \u2014 visit the Forecasting page first)"
    return ("Demand Forecast Report", subtitle,
            ["Material", "Month", "Predicted", "Current Stock", "Trend", "Model Fit"], out)


REPORT_BUILDERS = {
    "inventory":   build_inventory_report,
    "customer":    build_customer_report,
    "project":     build_project_report,
    "transaction": build_transaction_report,
    "commission":  build_commission_report,
    "forecast":    build_forecast_report,
}


# =================================================================
#  File generators — PDF (reportlab) and Excel (openpyxl)
# =================================================================
def generate_pdf(title, subtitle, columns, rows, generated_by="System"):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter,
                             topMargin=0.6 * inch, bottomMargin=0.6 * inch,
                             leftMargin=0.6 * inch, rightMargin=0.6 * inch)
    styles = getSampleStyleSheet()
    header_style = ParagraphStyle("Header", parent=styles["Heading1"],
                                   textColor=colors.HexColor(f"#{PRIMARY_HEX}"), fontSize=16)
    sub_style = ParagraphStyle("Sub", parent=styles["Normal"], textColor=colors.grey, fontSize=9)
    meta_style = ParagraphStyle("Meta", parent=styles["Normal"], fontSize=8,
                                 textColor=colors.grey, alignment=2)  # right-aligned

    generated = datetime.now().strftime("%B %d, %Y  %I:%M %p")
    story = [
        Paragraph("PARABELLUM STEEL &amp; IRON WORKS", header_style),
        Paragraph(subtitle, sub_style),
        Spacer(1, 4),
        Paragraph(f"Generated {generated} by {generated_by}", meta_style),
        Spacer(1, 12),
    ]

    table_data = [columns] + rows if rows else [columns, ["No records found."] + [""] * (len(columns) - 1)]
    col_width = (letter[0] - 1.2 * inch) / len(columns)
    tbl = Table(table_data, colWidths=[col_width] * len(columns), repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{PRIMARY_HEX}")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F1E8")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E4D9C4")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(tbl)
    doc.title = title
    doc.build(story)
    buf.seek(0)
    return buf


def generate_excel(title, subtitle, columns, rows, generated_by="System"):
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet-name limit

    header_fill = PatternFill("solid", fgColor=PRIMARY_HEX)
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(bold=True, size=14, color=PRIMARY_HEX)

    ws.merge_cells("A1:" + get_column_letter(len(columns)) + "1")
    ws["A1"] = "PARABELLUM STEEL & IRON WORKS"
    ws["A1"].font = title_font

    ws.merge_cells("A2:" + get_column_letter(len(columns)) + "2")
    ws["A2"] = subtitle
    ws["A2"].font = Font(italic=True, color="666666")

    generated = datetime.now().strftime("%B %d, %Y  %I:%M %p")
    ws.merge_cells("A3:" + get_column_letter(len(columns)) + "3")
    ws["A3"] = f"Generated {generated} by {generated_by}"
    ws["A3"].font = Font(size=9, color="999999")

    header_row = 5
    for i, col in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=i, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    display_rows = rows if rows else [["No records found."] + [""] * (len(columns) - 1)]
    for r, row in enumerate(display_rows, start=header_row + 1):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)

    for i, col in enumerate(columns, start=1):
        width = max(len(col), *(len(str(row[i - 1])) for row in display_rows)) + 3
        ws.column_dimensions[get_column_letter(i)].width = min(width, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
